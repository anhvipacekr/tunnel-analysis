"""
exporter.py - Export tunnel analysis results to CSV, Excel, and PDF.
Follows Leica Cyclone 3DR output conventions.
"""
from .common import *
from .models import PipelineContext, SectionGeometry
from pathlib import Path
from datetime import datetime


class TunnelExporter:
    """Export analysis results in industry-standard formats."""

    # Thresholds per PDF proposal
    THRESHOLDS = {
        "crown_settlement_mm":    {"caution": 10.0, "critical": 25.0},
        "lateral_convergence_mm": {"caution": 15.0, "critical": 30.0},
        "ovality_mean_pct":       {"caution":  0.5, "critical":  1.0},
        "eccentricity_mean_mm":   {"caution": 10.0, "critical": 25.0},
    }

    # ------------------------------------------------------------------ CSV --
    def export_csv(self, context: PipelineContext, out_path: str) -> str:
        """Export section-by-section parameters to CSV."""
        rows = self._build_rows(context)
        if not rows:
            raise RuntimeError("No section data to export. Run parameter extraction first.")
        import csv
        path = Path(out_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", newline="", encoding="utf-8-sig") as fh:
            writer = csv.DictWriter(fh, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)
        return str(path)

    # ---------------------------------------------------------------- Excel --
    def export_excel(self, context: PipelineContext, out_path: str,
                     project_name: str = "Tunnel Analysis",
                     engineer: str = "CBNU Smart Structure Lab") -> str:
        """Export full analysis report to Excel with multiple sheets."""
        try:
            import openpyxl
            from openpyxl.styles import (Font, PatternFill, Alignment,
                                          Border, Side, numbers)
            from openpyxl.chart import LineChart, Reference
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError("openpyxl required: pip install openpyxl")

        rows = self._build_rows(context)
        if not rows:
            raise RuntimeError("No section data to export.")

        wb = openpyxl.Workbook()

        # ---- Sheet 1: Cover ----
        ws_cover = wb.active
        ws_cover.title = "Cover"
        ws_cover.column_dimensions["A"].width = 30
        ws_cover.column_dimensions["B"].width = 40
        title_font = Font(name="Arial", size=16, bold=True, color="0F4C81")
        sub_font   = Font(name="Arial", size=11, color="475569")
        val_font   = Font(name="Arial", size=11, bold=True)
        ws_cover["A1"] = "TUNNEL STRUCTURAL HEALTH MONITORING"
        ws_cover["A1"].font = title_font
        ws_cover.merge_cells("A1:B1")
        ws_cover["A2"] = "LiDAR Point Cloud Analysis Report"
        ws_cover["A2"].font = sub_font
        ws_cover.merge_cells("A2:B2")
        ws_cover["A4"] = "Project:";    ws_cover["B4"] = project_name
        ws_cover["A5"] = "Engineer:";   ws_cover["B5"] = engineer
        ws_cover["A6"] = "Date:";       ws_cover["B6"] = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws_cover["A7"] = "Software:";   ws_cover["B7"] = "Tunnel Analysis v4.0 - CBNU"
        scan = context.active_scan
        ws_cover["A9"]  = "Scan file:"; ws_cover["B9"]  = scan.path if scan else "N/A"
        ws_cover["A10"] = "Points:";    ws_cover["B10"] = f"{len(scan.points):,}" if scan else "N/A"
        ws_cover["A11"] = "Sections:";  ws_cover["B11"] = len(rows)
        ws_cover["A12"] = "Profile:";   ws_cover["B12"] = context.tunnel_profile
        for row in ws_cover.iter_rows(min_row=4, max_row=12):
            for cell in row:
                if cell.column == 1:
                    cell.font = sub_font
                else:
                    cell.font = val_font

        # ---- Sheet 2: Section Data ----
        ws_data = wb.create_sheet("Section Data")
        headers = list(rows[0].keys())
        header_fill = PatternFill("solid", fgColor="0F4C81")
        header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        thin = Side(style="thin", color="CBD5E1")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col, h in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = border
            ws_data.column_dimensions[get_column_letter(col)].width = max(14, len(h) + 2)

        caution_fill  = PatternFill("solid", fgColor="FEF3C7")
        critical_fill = PatternFill("solid", fgColor="FEE2E2")
        ok_fill       = PatternFill("solid", fgColor="D1FAE5")
        for r, row in enumerate(rows, 2):
            for col, (key, val) in enumerate(row.items(), 1):
                cell = ws_data.cell(row=r, column=col, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
                if isinstance(val, float):
                    cell.number_format = "0.000"
                    status = self._status(key, val)
                    if status == "critical":
                        cell.fill = critical_fill
                    elif status == "caution":
                        cell.fill = caution_fill
                    elif status == "ok":
                        cell.fill = ok_fill
        ws_data.freeze_panes = "A2"
        ws_data.auto_filter.ref = ws_data.dimensions

        # ---- Sheet 3: Summary Stats ----
        ws_sum = wb.create_sheet("Summary")
        ws_sum["A1"] = "Parameter Summary"
        ws_sum["A1"].font = Font(name="Arial", size=13, bold=True, color="0F4C81")
        ws_sum.merge_cells("A1:E1")
        sum_headers = ["Parameter", "Min", "Max", "Mean", "Status"]
        for col, h in enumerate(sum_headers, 1):
            cell = ws_sum.cell(row=2, column=col, value=h)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws_sum.column_dimensions[get_column_letter(col)].width = 22
        params = ["chainage_m", "crown_settlement_mm", "lateral_convergence_mm",
                  "ovality_pct", "eccentricity_mm", "radius_fit_m",
                  "clearance_violation"]
        for r, param in enumerate(params, 3):
            vals = [row.get(param) for row in rows if isinstance(row.get(param), (int, float))]
            if not vals:
                continue
            arr = np.array(vals, dtype=np.float64)
            status = self._status(param, float(np.nanmax(arr)))
            fill = critical_fill if status == "critical" else caution_fill if status == "caution" else ok_fill
            for col, val in enumerate([param, float(np.nanmin(arr)), float(np.nanmax(arr)),
                                        float(np.nanmean(arr)), status.upper()], 1):
                cell = ws_sum.cell(row=r, column=col, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
                if isinstance(val, float):
                    cell.number_format = "0.000"
                if col >= 2:
                    cell.fill = fill

        # ---- Sheet 4: Charts ----
        ws_chart = wb.create_sheet("Charts")
        # Write chainage + key params for charting
        chart_headers = ["Chainage (m)", "Settlement (mm)", "Convergence (mm)", "Ovality (%)"]
        for col, h in enumerate(chart_headers, 1):
            ws_chart.cell(row=1, column=col, value=h).font = header_font
        for r, row in enumerate(rows, 2):
            ws_chart.cell(row=r, column=1, value=row.get("chainage_m", r-1))
            ws_chart.cell(row=r, column=2, value=row.get("crown_settlement_mm", None))
            ws_chart.cell(row=r, column=3, value=row.get("lateral_convergence_mm", None))
            ws_chart.cell(row=r, column=4, value=row.get("ovality_pct", None))

        n = len(rows) + 1
        for col_idx, (title, col_letter) in enumerate([
            ("Crown Settlement (mm)", "B"),
            ("Horizontal Convergence (mm)", "C"),
            ("Ovality (%)", "D"),
        ], 1):
            chart = LineChart()
            chart.title = title
            chart.style = 10
            chart.y_axis.title = title
            chart.x_axis.title = "Chainage (m)"
            chart.width = 20; chart.height = 10
            data = Reference(ws_chart, min_col=col_idx+1, min_row=1, max_row=n)
            cats = Reference(ws_chart, min_col=1, min_row=2, max_row=n)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.series[0].graphicalProperties.line.solidFill = "1D4ED8"
            ws_chart.add_chart(chart, f"F{1 + (col_idx-1)*18}")

        # ---- Sheet 5: Warnings ----
        ws_warn = wb.create_sheet("Warnings")
        ws_warn["A1"] = "Structural Warning Report"
        ws_warn["A1"].font = Font(name="Arial", size=13, bold=True, color="DC2626")
        ws_warn.merge_cells("A1:F1")
        warn_headers = ["Chainage (m)", "Parameter", "Value", "Threshold", "Status", "Action"]
        for col, h in enumerate(warn_headers, 1):
            cell = ws_warn.cell(row=2, column=col, value=h)
            cell.font = header_font; cell.fill = header_fill
            ws_warn.column_dimensions[get_column_letter(col)].width = 20
        warn_row = 3
        for row in rows:
            ch = row.get("chainage_m", 0)
            for param, thr in self.THRESHOLDS.items():
                val = row.get(param)
                if not isinstance(val, (int, float)) or not np.isfinite(val):
                    continue
                if val >= thr["critical"]:
                    status = "CRITICAL"
                    action = "Immediate inspection required"
                    fill = critical_fill
                elif val >= thr["caution"]:
                    status = "CAUTION"
                    action = "Schedule inspection within 30 days"
                    fill = caution_fill
                else:
                    continue
                for col, v in enumerate([ch, param, round(val, 3),
                                          thr["critical"], status, action], 1):
                    cell = ws_warn.cell(row=warn_row, column=col, value=v)
                    cell.fill = fill; cell.border = border
                    cell.alignment = Alignment(horizontal="center")
                warn_row += 1
        if warn_row == 3:
            ws_warn["A3"] = "No structural warnings detected."
            ws_warn["A3"].font = Font(color="047857", bold=True)

        path = Path(out_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))
        return str(path)

    # ---------------------------------------------------------------- helpers --
    def _build_rows(self, context: PipelineContext) -> List[Dict]:
        rows = []
        params = context.parameters
        if context.sections:
            for sec in context.sections:
                row = {
                    "chainage_m":             round(sec.chainage, 3),
                    "crown_settlement_mm":    round(sec.H1 * 1e3, 3) if np.isfinite(sec.H1) else None,
                    "lateral_convergence_mm": round(sec.W1 * 1e3, 3) if np.isfinite(sec.W1) else None,
                    "ovality_pct":            round(sec.ovality, 3) if np.isfinite(sec.ovality) else None,
                    "eccentricity_mm":        round(sec.eccentricity, 3) if np.isfinite(sec.eccentricity) else None,
                    "radius_fit_m":           round(sec.radius_fit, 4) if np.isfinite(sec.radius_fit) else None,
                    "wall_angle_L_deg":       round(sec.wall_angle_L, 2) if np.isfinite(sec.wall_angle_L) else None,
                    "wall_angle_R_deg":       round(sec.wall_angle_R, 2) if np.isfinite(sec.wall_angle_R) else None,
                    "clearance_violation":    int(sec.clearance_violation),
                    "min_clearance_dist_m":   round(sec.min_clearance_dist, 4) if np.isfinite(sec.min_clearance_dist) else None,
                }
                rows.append(row)
        elif params:
            rows.append({k: round(v, 4) if isinstance(v, float) else v
                         for k, v in params.items()})
        return rows

    def _status(self, key: str, val: float) -> str:
        thr = self.THRESHOLDS.get(key)
        if thr is None or not np.isfinite(val):
            return "n/a"
        if val >= thr["critical"]:
            return "critical"
        if val >= thr["caution"]:
            return "caution"
        return "ok"
